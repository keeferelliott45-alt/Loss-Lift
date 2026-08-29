"""Stage 7 — export (spec section 5).

One workbook, four sheets:

* **Claims** — the table, in the chosen column order, with cells that carry a
  finding shaded amber and vision-extracted cells marked.
* **Loss Summary** — claims by policy term, each checked against the subtotal
  the carrier printed for that term.
* **Exceptions** — every finding, errors first, with expected, actual and delta.
* **Source Info** — filename, hash, valuation date, extraction method,
  timestamp and reconciliation status, so the workbook can be audited back to
  the document it came from.

``build_account_workbook`` writes a different shape for a whole account —
several runs for one insured merged, with each claim's development between
valuations and a sheet naming every file the history came from.

The redaction toggle drops claimant names and loss descriptions (spec
section 9).
"""

from __future__ import annotations

import io
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.summary import summarise_by_period

if TYPE_CHECKING:  # imported for typing only; core.account imports nothing here
    from core.account import AccountRollup
from core.schema import (
    DATE_FIELDS,
    MONEY_FIELDS,
    REDACTED_FIELDS,
    Claim,
    DocumentStatus,
    LossRunDocument,
    ReconciliationResult,
    Severity,
    SourceMethod,
)

#: Column-order templates offered on the export screen.
EXPORT_TEMPLATES: dict[str, tuple[str, ...]] = {
    "Underwriting standard": (
        "claim_number",
        "date_of_loss",
        "date_reported",
        "claim_status",
        "cause_of_loss",
        "loss_description",
        "paid_total",
        "reserve_total",
        "recovery_total",
        "incurred_total",
    ),
    "Full detail": (
        "claim_number",
        "date_of_loss",
        "date_reported",
        "claim_status",
        "claimant_name",
        "loss_description",
        "cause_of_loss",
        "paid_indemnity",
        "paid_medical",
        "paid_expense",
        "paid_total",
        "reserve_indemnity",
        "reserve_medical",
        "reserve_expense",
        "reserve_total",
        "recovery_total",
        "incurred_total",
        "litigation_flag",
    ),
    "Workers comp": (
        "claim_number",
        "date_of_loss",
        "date_reported",
        "claim_status",
        "claimant_name",
        "paid_indemnity",
        "paid_medical",
        "paid_expense",
        "paid_total",
        "reserve_indemnity",
        "reserve_medical",
        "reserve_expense",
        "reserve_total",
        "incurred_total",
    ),
    "Claim numbers and incurred": (
        "claim_number",
        "date_of_loss",
        "incurred_total",
    ),
}

DEFAULT_TEMPLATE = "Underwriting standard"

COLUMN_TITLES: dict[str, str] = {
    "claim_number": "Claim number",
    "date_of_loss": "Date of loss",
    "date_reported": "Date reported",
    "claim_status": "Status",
    "claimant_name": "Claimant",
    "loss_description": "Loss description",
    "cause_of_loss": "Cause of loss",
    "paid_indemnity": "Paid indemnity",
    "paid_medical": "Paid medical",
    "paid_expense": "Paid expense",
    "paid_total": "Paid total",
    "reserve_indemnity": "Reserve indemnity",
    "reserve_medical": "Reserve medical",
    "reserve_expense": "Reserve expense",
    "reserve_total": "Reserve total",
    "recovery_total": "Recovery",
    "incurred_total": "Total incurred",
    "litigation_flag": "In suit",
    "source_page": "Source page",
    "source_method": "Extraction",
}

MONEY_FORMAT = '#,##0.00;[Red](#,##0.00)'
DATE_FORMAT = "yyyy-mm-dd"

#: Default cut for the Large Loss sheet. Configurable per export.
LARGE_LOSS_THRESHOLD = Decimal("25000")

#: Document-level facts written onto every claim row, so one row describes
#: itself without its header — what makes a merged multi-carrier sheet usable.
_DOCUMENT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Carrier", "carrier"),
    ("Named insured", "named_insured"),
    ("Policy number", "policy_number"),
    ("Policy term start", "policy_period_start"),
    ("Policy term end", "policy_period_end"),
    ("Line of business", "line_of_business"),
    ("Valuation date", "valuation_date"),
)


def _document_values(document: LossRunDocument) -> list[Any]:
    values: list[Any] = []
    for _, attribute in _DOCUMENT_COLUMNS:
        value = getattr(document, attribute, None)
        values.append(value.value if hasattr(value, "value") else value)
    return values


_HEADER_FILL = PatternFill("solid", fgColor="1F2933")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_FINDING_FILL = PatternFill("solid", fgColor="FDE8C8")   # amber
_ERROR_FILL = PatternFill("solid", fgColor="F8D7DA")     # red
_VISION_FONT = Font(italic=True, color="6B5B00")


def _title(field_name: str) -> str:
    return COLUMN_TITLES.get(field_name, field_name.replace("_", " ").capitalize())


def resolve_columns(
    template: str | Sequence[str] = DEFAULT_TEMPLATE,
    *,
    redact: bool = False,
    include_provenance: bool = True,
) -> list[str]:
    """The column order for the Claims sheet."""
    if isinstance(template, str):
        columns = list(EXPORT_TEMPLATES.get(template, EXPORT_TEMPLATES[DEFAULT_TEMPLATE]))
    else:
        columns = list(template)
    if redact:
        columns = [name for name in columns if name not in REDACTED_FIELDS]
    if include_provenance:
        columns += ["source_page", "source_method"]
    return columns


def _cell_value(claim: Claim, field_name: str) -> Any:
    value = getattr(claim, field_name, None)
    if value is None:
        return None
    if isinstance(value, Decimal):
        # openpyxl has no Decimal type; float here is presentation only and
        # never feeds a reconciliation.
        return float(value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, date):
        return value
    if hasattr(value, "value"):
        return value.value
    return str(value)


def _findings_index(
    result: ReconciliationResult | None,
) -> dict[tuple[str, str], Severity]:
    """Which (claim, field) pairs carry a finding, and how bad."""
    index: dict[tuple[str, str], Severity] = {}
    if result is None:
        return index
    for finding in result.findings:
        if not finding.claim_number or not finding.field:
            continue
        key = (finding.claim_number, finding.field)
        current = index.get(key)
        if current is None or finding.severity is Severity.ERROR:
            index[key] = finding.severity
    return index


def _autosize(sheet: Worksheet, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width, 10), 52)


def _write_claims_sheet(
    sheet: Worksheet,
    document: LossRunDocument,
    columns: Sequence[str],
    result: ReconciliationResult | None,
) -> None:
    findings = _findings_index(result)
    widths: dict[int, int] = {}

    titles = [_title(name) for name in columns] + [
        title for title, _ in _DOCUMENT_COLUMNS
    ]
    for column_index, title in enumerate(titles, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths[column_index] = len(title) + 2

    document_values = _document_values(document)

    for row_index, claim in enumerate(document.claims, start=2):
        for column_index, field_name in enumerate(columns, start=1):
            value = _cell_value(claim, field_name)
            cell = sheet.cell(row=row_index, column=column_index, value=value)

            if field_name in MONEY_FIELDS:
                cell.number_format = MONEY_FORMAT
            elif field_name in DATE_FIELDS:
                cell.number_format = DATE_FORMAT

            severity = findings.get((claim.claim_number, field_name))
            if severity is Severity.ERROR:
                cell.fill = _ERROR_FILL
            elif severity is not None:
                cell.fill = _FINDING_FILL
            elif claim.field_issues.get(field_name):
                cell.fill = _FINDING_FILL

            if claim.source_method is SourceMethod.VISION:
                cell.font = _VISION_FONT

            if value is not None:
                widths[column_index] = max(widths[column_index], len(str(value)) + 2)

        for offset, value in enumerate(document_values):
            column_index = len(columns) + 1 + offset
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if isinstance(value, date):
                cell.number_format = DATE_FORMAT
            if value is not None:
                widths[column_index] = max(widths[column_index], len(str(value)) + 2)

    sheet.freeze_panes = "A2"
    if document.claims:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{len(document.claims) + 1}"
        )
    _autosize(sheet, widths)


def _write_exceptions_sheet(
    sheet: Worksheet, result: ReconciliationResult | None
) -> None:
    headers = ["Rule", "Severity", "Claim number", "Field", "What happened",
               "Expected", "Actual", "Delta", "Page"]
    widths: dict[int, int] = {}
    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        widths[column_index] = len(title) + 2

    findings = result.findings if result else []
    for row_index, finding in enumerate(findings, start=2):
        values = [
            finding.rule_id,
            finding.severity.value,
            finding.claim_number or "",
            _title(finding.field) if finding.field else "",
            finding.message,
            float(finding.expected) if isinstance(finding.expected, Decimal) else finding.expected,
            float(finding.actual) if isinstance(finding.actual, Decimal) else finding.actual,
            float(finding.delta) if finding.delta is not None else None,
            finding.page,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if column_index in (6, 7, 8) and isinstance(value, float):
                cell.number_format = MONEY_FORMAT
            if finding.severity is Severity.ERROR:
                cell.fill = _ERROR_FILL
            elif finding.severity is Severity.WARN:
                cell.fill = _FINDING_FILL
            if value is not None:
                widths[column_index] = max(widths[column_index], len(str(value)) + 2)

    if not findings:
        sheet.cell(row=2, column=1, value="No exceptions. Every check passed.")

    sheet.freeze_panes = "A2"
    _autosize(sheet, widths)


def _write_summary_sheet(sheet: Worksheet, document: LossRunDocument) -> None:
    """Claims by policy term — the loss history a submission asks for.

    "Ties to carrier" is the column that matters: it says, per term, whether
    these numbers match the subtotal the carrier printed for that term.
    """
    headers = ["Policy term", "Claims", "Open", "Closed", "Paid", "Reserves",
               "Recoveries", "Incurred", "Frequency", "Severity",
               "Largest loss", "Carrier printed", "Ties to carrier"]
    widths: dict[int, int] = {}
    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        widths[column_index] = len(title) + 2

    periods = summarise_by_period(document)
    for row_index, period in enumerate(periods, start=2):
        ties = period.ties()
        printed = period.printed_totals.get("incurred_total")
        values = [
            period.label,
            period.claims,
            period.open_claims,
            period.closed_claims,
            float(period.totals["paid_total"]),
            float(period.totals["reserve_total"]),
            float(period.totals["recovery_total"]),
            float(period.totals["incurred_total"]),
            # Frequency is the claim count for the term; severity is the mean
            # incurred per claim. No exposure base is on a loss run, so neither
            # is rated per unit of payroll or revenue -- that is the broker's
            # own number to divide by.
            period.claims,
            (
                float(period.totals["incurred_total"] / period.claims)
                if period.claims
                else None
            ),
            float(period.largest_loss) if period.largest_loss is not None else None,
            float(printed) if printed is not None else None,
            "not printed" if ties is None else "yes" if ties else "no",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if isinstance(value, float):
                cell.number_format = MONEY_FORMAT
            if ties is False:
                cell.fill = _ERROR_FILL
            if value is not None:
                widths[column_index] = max(widths[column_index], len(str(value)) + 2)

    if not periods:
        sheet.cell(row=2, column=1, value="No claims to summarise.")

    sheet.freeze_panes = "A2"
    _autosize(sheet, widths)


def _write_large_loss_sheet(
    sheet: Worksheet,
    document: LossRunDocument,
    threshold: Decimal,
) -> None:
    """Claims at or above the threshold, worst first.

    The shock losses drive the price, so they get their own sheet rather than
    being found by sorting the detail tab.
    """
    headers = ["Claim number", "Date of loss", "Status", "Cause of loss",
               "Paid", "Reserve", "Recovery", "Incurred", "Source page"]
    widths = _header_row(sheet, headers)

    large = sorted(
        (c for c in document.claims
         if c.incurred_total is not None and c.incurred_total >= threshold),
        key=lambda c: c.incurred_total,
        reverse=True,
    )
    for row_index, claim in enumerate(large, start=2):
        _fill_row(sheet, row_index, [
            claim.claim_number,
            claim.date_of_loss,
            claim.claim_status.value if claim.claim_status else None,
            claim.cause_of_loss,
            _float(claim.paid_total),
            _float(claim.reserve_total),
            _float(claim.recovery_total),
            _float(claim.incurred_total),
            claim.source_page,
        ], widths)

    if not large:
        sheet.cell(row=2, column=1,
                   value=f"No claim reaches {threshold:,.0f}.")
    sheet.freeze_panes = "A2"
    _autosize(sheet, widths)


def _write_source_sheet(
    sheet: Worksheet,
    document: LossRunDocument,
    result: ReconciliationResult | None,
    *,
    redacted: bool,
    template: str,
) -> None:
    status = result.status if result else DocumentStatus.CLEAN
    error_count = len(result.errors) if result else 0
    warn_count = len(result.warnings) if result else 0

    rows: list[tuple[str, Any]] = [
        ("Source file", document.source_filename),
        ("SHA-256", document.file_sha256),
        ("Carrier", document.carrier or "not found"),
        ("Named insured", document.named_insured or "not found"),
        ("Policy number", document.policy_number or "not found"),
        ("Policy period", _period(document)),
        ("Line of business", document.line_of_business.value if document.line_of_business else "not found"),
        ("Valuation date", document.valuation_date or "MISSING — see exceptions"),
        ("Currency", document.currency),
        ("Number format", "European (1.234,56)" if document.locale_hint == "eu" else "US (1,234.56)"),
        ("Number format proven", "yes" if document.locale_confident else "no — assumed"),
        ("Date order", document.date_order or "unknown"),
        ("Date order proven", "yes" if document.date_order_confident else "no — assumed"),
        ("Recoveries printed as", document.recovery_convention_label),
        ("Pages", document.page_count),
        ("Extraction method", document.extraction_method.value),
        ("Scanned pages", ", ".join(str(p) for p in document.scanned_pages) or "none"),
        ("Carrier profile", document.profile_name or "none saved"),
        ("Claims extracted", len(document.claims)),
        ("Claim count printed on document", document.printed_claim_count if document.printed_claim_count is not None else "not printed"),
        ("Reconciliation status", "Reconciled" if status is DocumentStatus.CLEAN else "Needs review"),
        ("Errors", error_count),
        ("Warnings", warn_count),
        ("Claimant data redacted", "yes" if redacted else "no"),
        ("Column template", template if isinstance(template, str) else "custom"),
        ("Exported at", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]

    for row_index, (label, value) in enumerate(rows, start=1):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        label_cell.font = Font(bold=True)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        if isinstance(value, date):
            value_cell.number_format = DATE_FORMAT

    blank = len(rows) + 2
    sheet.cell(row=blank, column=1, value="Printed totals vs extracted").font = Font(bold=True)
    header_row = blank + 1
    for column_index, title in enumerate(
        ["Column", "Printed on document", "Extracted total", "Difference"], start=1
    ):
        cell = sheet.cell(row=header_row, column=column_index, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_index = header_row + 1
    for field_name in MONEY_FIELDS:
        printed = document.printed_totals.get(field_name)
        if printed is None:
            continue
        extracted = document.column_total(field_name)
        sheet.cell(row=row_index, column=1, value=_title(field_name))
        for column_index, amount in enumerate(
            (printed, extracted, extracted - printed), start=2
        ):
            cell = sheet.cell(row=row_index, column=column_index, value=float(amount))
            cell.number_format = MONEY_FORMAT
        row_index += 1
    if row_index == header_row + 1:
        sheet.cell(row=row_index, column=1, value="The document printed no column totals.")

    _autosize(sheet, {1: 34, 2: 46, 3: 20, 4: 18})


def _period(document: LossRunDocument) -> str:
    start, end = document.policy_period_start, document.policy_period_end
    if not start and not end:
        return "not found"
    return f"{start or '?'} to {end or '?'}"


def build_workbook(
    document: LossRunDocument,
    result: ReconciliationResult | None = None,
    *,
    template: str | Sequence[str] = DEFAULT_TEMPLATE,
    redact: bool = False,
    include_provenance: bool = True,
    large_loss_threshold: Decimal = LARGE_LOSS_THRESHOLD,
) -> Workbook:
    """Build the workbook: claims, loss summary, exceptions and provenance."""
    columns = resolve_columns(
        template, redact=redact, include_provenance=include_provenance
    )
    workbook = Workbook()
    claims_sheet = workbook.active
    claims_sheet.title = "Claim Detail"
    _write_claims_sheet(claims_sheet, document, columns, result)
    _write_summary_sheet(workbook.create_sheet("Loss Summary"), document)
    _write_large_loss_sheet(
        workbook.create_sheet("Large Loss"), document, large_loss_threshold
    )
    _write_exceptions_sheet(workbook.create_sheet("Exceptions"), result)
    _write_source_sheet(
        workbook.create_sheet("Source Info"),
        document,
        result,
        redacted=redact,
        template=template if isinstance(template, str) else "custom",
    )
    return workbook


def to_bytes(
    document: LossRunDocument,
    result: ReconciliationResult | None = None,
    **kwargs: Any,
) -> bytes:
    """The workbook as bytes, for a Streamlit download button."""
    buffer = io.BytesIO()
    build_workbook(document, result, **kwargs).save(buffer)
    return buffer.getvalue()


def _header_row(sheet: Worksheet, headers: Sequence[str]) -> dict[int, int]:
    widths: dict[int, int] = {}
    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        widths[column_index] = len(title) + 2
    return widths


def _fill_row(
    sheet: Worksheet, row_index: int, values: Sequence[Any], widths: dict[int, int]
) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        if isinstance(value, float):
            cell.number_format = MONEY_FORMAT
        elif isinstance(value, date):
            cell.number_format = DATE_FORMAT
        if value is not None:
            widths[column_index] = max(widths[column_index], len(str(value)) + 2)


def build_account_workbook(account: "AccountRollup", *, redact: bool = False) -> Workbook:
    """One insured's merged loss history.

    Deliberately not the single-document workbook pointed at merged claims: the
    provenance of a merged history is several files with several valuation
    dates, and Source Info can only name one. The Sources sheet names them all.
    """
    workbook = Workbook()

    claims_sheet = workbook.active
    claims_sheet.title = "Claims"
    headers = ["Claim number", "Date of loss", "Status", "Carrier", "Valued at",
               "Paid", "Reserve", "Recovery", "Incurred", "Development", "Runs seen in"]
    if not redact:
        headers.insert(3, "Claimant")
    widths = _header_row(claims_sheet, headers)

    for row_index, history in enumerate(account.histories, start=2):
        claim = history.current
        values: list[Any] = [
            history.claim_number,
            claim.date_of_loss,
            claim.claim_status.value if claim.claim_status else None,
        ]
        if not redact:
            values.append(claim.claimant_name)
        values += [
            ", ".join(history.carriers) or None,
            history.valued_at,
            _float(claim.paid_total),
            _float(claim.reserve_total),
            _float(claim.recovery_total),
            _float(claim.incurred_total),
            _float(history.development),
            len(history.appearances),
        ]
        _fill_row(claims_sheet, row_index, values, widths)
        if history.development and history.development > 0:
            claims_sheet.cell(row=row_index, column=len(headers) - 1).fill = _FINDING_FILL
    claims_sheet.freeze_panes = "A2"
    _autosize(claims_sheet, widths)

    summary_sheet = workbook.create_sheet("Loss Summary")
    widths = _header_row(summary_sheet, ["Policy term", "Claims", "Open", "Closed",
                                         "Paid", "Reserves", "Recoveries", "Incurred",
                                         "Largest loss"])
    for row_index, period in enumerate(account.periods, start=2):
        _fill_row(summary_sheet, row_index, [
            period.label, period.claims, period.open_claims, period.closed_claims,
            float(period.totals["paid_total"]),
            float(period.totals["reserve_total"]),
            float(period.totals["recovery_total"]),
            float(period.totals["incurred_total"]),
            _float(period.largest_loss),
        ], widths)
    summary_sheet.freeze_panes = "A2"
    _autosize(summary_sheet, widths)

    sources_sheet = workbook.create_sheet("Sources")
    widths = _header_row(sources_sheet, ["File", "Carrier", "Valuation date",
                                         "Claims", "SHA-256"])
    for row_index, document in enumerate(account.documents, start=2):
        _fill_row(sources_sheet, row_index, [
            document.source_filename,
            document.carrier,
            document.valuation_date,
            len(document.claims),
            document.file_sha256,
        ], widths)
    _autosize(sources_sheet, widths)
    return workbook


def account_to_bytes(account: "AccountRollup", **kwargs: Any) -> bytes:
    """The account workbook as bytes, for a Streamlit download button."""
    buffer = io.BytesIO()
    build_account_workbook(account, **kwargs).save(buffer)
    return buffer.getvalue()


def _float(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def write_xlsx(
    document: LossRunDocument,
    path: str | Path,
    result: ReconciliationResult | None = None,
    **kwargs: Any,
) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(document, result, **kwargs).save(target)
    return target


def suggested_filename(document: LossRunDocument) -> str:
    """A filename an account manager can find again."""
    stem = Path(document.source_filename).stem or "loss-run"
    parts = [part for part in (document.named_insured, stem) if part]
    label = " - ".join(parts)[:80]
    valuation = document.valuation_date.isoformat() if document.valuation_date else "no-valuation-date"
    safe = "".join(ch if ch.isalnum() or ch in " -_." else "-" for ch in label).strip()
    return f"{safe} {valuation}.xlsx".replace("  ", " ")
