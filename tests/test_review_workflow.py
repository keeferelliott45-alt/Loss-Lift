"""Findings a reviewer can act on, without review being mistaken for proof.

LossLift holds three kinds of fact and this file exists to stop one wearing
another's clothes: what the carrier printed, what LossLift made of it, and what
a person decided afterwards.

The temptation the tests guard against is the tidy one. A screen full of
resolved findings looks like a document in good order, and it is not: the
engine is never told that anyone reviewed anything, so a document that failed
R-04 before the review fails it after. Only a changed value can move a
reconciliation result, and then only because every rule ran again over it.
"""

from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from core.export import to_bytes
from core.pipeline import (
    apply_edits,
    rerun_reconciliation,
    resolve_finding,
    review_columns,
    run_pipeline,
    to_records,
)
from core.review import (
    EXTRACTION_RULES,
    FINANCIAL_RULES,
    ReviewAction,
    bucket_of,
    finding_key,
    summarise_review,
)
from core.evidence import EvidenceKind, claim_evidence, finding_evidence
from core.schema import DocumentStatus, SourceMethod


@pytest.fixture()
def broken(golden_dir):
    """A document with a real arithmetic error planted in it."""
    return run_pipeline(golden_dir / "arithmetic_error.pdf", use_vision=False)


def _finding(result, rule_id: str):
    return next(f for f in result.reconciliation.findings if f.rule_id == rule_id)


# --------------------------------------------------------------------------
# The four readings stay four
# --------------------------------------------------------------------------


def test_the_three_questions_are_asked_separately(broken):
    summary = summarise_review(broken.reconciliation.findings, broken.document.review_log)
    assert summary.total == len(broken.reconciliation.findings)
    assert summary.financial.total + summary.extraction.total + summary.underwriting.total == summary.total
    assert not FINANCIAL_RULES & EXTRACTION_RULES


def test_reviewing_never_makes_a_document_reconcile(broken):
    """The load-bearing one. Resolve everything; the figures still disagree."""
    result = broken
    assert not summarise_review(result.reconciliation.findings, result.document.review_log).financial.passes

    for finding in list(result.reconciliation.findings):
        result = resolve_finding(result, finding, ReviewAction.DISMISSED, note="looked")

    summary = summarise_review(result.reconciliation.findings, result.document.review_log)
    assert summary.fully_reviewed
    assert not summary.financial.passes
    assert summary.headline() != "reconciled"
    assert result.reconciliation.status is DocumentStatus.NEEDS_REVIEW


def test_a_dismissed_flag_changes_no_extracted_data(broken):
    before = broken.document.model_dump_json(exclude={"review_log"})
    flag = next(
        f for f in broken.reconciliation.findings if bucket_of(f) == "underwriting"
    )
    result = resolve_finding(broken, flag, ReviewAction.DISMISSED)
    assert result.document.model_dump_json(exclude={"review_log"}) == before


def test_confirming_leaves_the_reconciliation_where_it_was(broken):
    status = broken.reconciliation.status
    findings = list(broken.reconciliation.findings)
    result = resolve_finding(broken, findings[0], ReviewAction.CONFIRMED, note="genuine")
    assert result.reconciliation.status is status
    assert [f.rule_id for f in result.reconciliation.findings] == [f.rule_id for f in findings]


# --------------------------------------------------------------------------
# The finding outlives the decision
# --------------------------------------------------------------------------


def test_the_original_finding_survives_being_resolved(broken):
    finding = _finding(broken, "R-01")
    result = resolve_finding(broken, finding, ReviewAction.DISMISSED, note="known")
    assert any(f.rule_id == "R-01" for f in result.reconciliation.findings)
    entry = result.document.review_log.latest_for(finding_key(finding))
    assert entry.message == finding.message
    assert entry.severity == finding.severity.value


def test_a_second_decision_appends_rather_than_rewrites(broken):
    finding = _finding(broken, "R-01")
    result = resolve_finding(broken, finding, ReviewAction.DISMISSED, note="first")
    result = resolve_finding(result, finding, ReviewAction.CONFIRMED, note="second")
    entries = [e for e in result.document.review_log.entries if e.key == finding_key(finding)]
    assert [e.note for e in entries] == ["first", "second"]
    assert result.document.review_log.action_for(finding) is ReviewAction.CONFIRMED


# --------------------------------------------------------------------------
# Corrections
# --------------------------------------------------------------------------


def test_a_correction_can_turn_a_failure_into_a_pass(broken):
    """The engine runs again over the corrected value, and agrees."""
    assert broken.reconciliation.status is DocumentStatus.NEEDS_REVIEW
    finding = _finding(broken, "R-01")
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED,
        corrected_value=str(finding.expected), note="misread on the page",
    )
    assert result.reconciliation.status is DocumentStatus.CLEAN
    assert not any(f.rule_id == "R-01" for f in result.reconciliation.findings)


def test_a_correction_can_leave_the_failure_standing(broken):
    """Correcting a value is not asserting that it fixed anything."""
    finding = _finding(broken, "R-01")
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED,
        corrected_value="1.00", note="still wrong",
    )
    assert result.reconciliation.status is DocumentStatus.NEEDS_REVIEW
    assert any(f.rule_id == "R-01" for f in result.reconciliation.findings)
    entry = result.document.review_log.entries[-1]
    assert entry.status_before == entry.status_after


def test_a_correction_records_what_was_there_before(broken):
    finding = _finding(broken, "R-01")
    was = next(
        c for c in broken.document.claims if c.claim_number == finding.claim_number
    ).incurred_total
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED, corrected_value=str(finding.expected)
    )
    entry = result.document.review_log.entries[-1]
    assert entry.before == f"{was:f}"
    assert entry.after == f"{Decimal(str(finding.expected)):f}"
    assert entry.changed_a_value


def test_a_correction_leaves_the_extracted_value_recoverable(broken):
    finding = _finding(broken, "R-01")
    was = next(
        c for c in broken.document.claims if c.claim_number == finding.claim_number
    ).incurred_total
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED, corrected_value=str(finding.expected)
    )
    claim = next(
        c for c in result.document.claims if c.claim_number == finding.claim_number
    )
    assert claim.original_of("incurred_total") == f"{was:f}"
    assert claim.incurred_total != was


# --------------------------------------------------------------------------
# Provenance under correction
# --------------------------------------------------------------------------


def test_a_corrected_field_is_manual_and_its_neighbours_are_not(broken):
    finding = _finding(broken, "R-01")
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED, corrected_value=str(finding.expected)
    )
    claim = next(
        c for c in result.document.claims if c.claim_number == finding.claim_number
    )
    assert claim.provenance_of("incurred_total") is SourceMethod.MANUAL
    assert claim.provenance_of("paid_total") is SourceMethod.DIGITAL


def test_evidence_survives_a_correction(broken):
    """Correcting a cell must not cost the row the line it was read from."""
    finding = _finding(broken, "R-01")
    before = claim_evidence(
        next(c for c in broken.document.claims if c.claim_number == finding.claim_number)
    )
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED, corrected_value=str(finding.expected)
    )
    claim = next(
        c for c in result.document.claims if c.claim_number == finding.claim_number
    )
    after = claim_evidence(claim, "paid_total")
    assert after.kind is EvidenceKind.REGION
    assert after.bbox == before.bbox
    assert claim_evidence(claim, "incurred_total").kind is EvidenceKind.TYPED


def test_a_document_wide_finding_is_given_no_claim_evidence(broken):
    """R-06 and its kind are about the document; there is no row to point at."""
    for finding in broken.reconciliation.findings:
        if finding.claim_number:
            continue
        evidence = finding_evidence(broken.document, finding)
        assert evidence.bbox is None
        assert evidence.kind in (EvidenceKind.PAGE, EvidenceKind.NONE)


# --------------------------------------------------------------------------
# Survival
# --------------------------------------------------------------------------


def test_review_history_survives_an_unrelated_edit(broken):
    """Editing the table rebuilds every claim; the log is not a claim.

    The edit adds its own line, because replacing a carrier's figure in the
    grid is the same act as correcting one against a finding. What must not
    happen is the earlier decision being rewritten or dropped on the way.
    """
    finding = _finding(broken, "R-01")
    result = resolve_finding(broken, finding, ReviewAction.CONFIRMED, note="kept")
    records = to_records(result.document, review_columns(result.document))
    records[0]["claimant_name"] = "Someone Else"
    updated = apply_edits(result.document, records)

    kept = updated.review_log.entries[0]
    assert kept.note == "kept"
    assert kept.action is ReviewAction.CONFIRMED
    assert kept.rule_id == "R-01"

    recorded = updated.review_log.entries[-1]
    assert recorded.action is ReviewAction.EDITED
    assert recorded.field == "claimant_name"
    assert (recorded.before, recorded.after) == ("Adams, Cole", "Someone Else")


def test_review_history_survives_a_reconciliation_rerun(broken):
    finding = _finding(broken, "R-01")
    result = resolve_finding(broken, finding, ReviewAction.DISMISSED, note="kept")
    rerun_reconciliation(result.document)
    assert result.document.review_log.action_for(finding) is ReviewAction.DISMISSED


def test_review_history_reaches_the_workbook(broken):
    finding = _finding(broken, "R-01")
    was = next(
        c for c in broken.document.claims if c.claim_number == finding.claim_number
    ).incurred_total
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED,
        corrected_value=str(finding.expected), note="misread on the page",
    )
    book = load_workbook(
        io.BytesIO(to_bytes(result.document, result.reconciliation))
    )
    assert "Review History" in book.sheetnames

    sheet = book["Review History"]
    headers = [cell.value for cell in sheet[1]]
    row = {name: sheet.cell(row=2, column=index + 1).value
           for index, name in enumerate(headers)}
    assert row["Decision"] == "corrected"
    assert row["Value before"] == f"{was:f}"
    assert row["Reviewer note"] == "misread on the page"
    # Named for what it carries: the document's standing, which review
    # progress moves and reconciliation is only part of.
    assert row["Document status before"] != row["Document status after"]
    assert row["Reviewer"] == "local reviewer"


def test_the_workbook_keeps_the_finding_and_the_decision_together(broken):
    finding = _finding(broken, "R-01")
    result = resolve_finding(broken, finding, ReviewAction.DISMISSED, note="known")
    sheet = load_workbook(
        io.BytesIO(to_bytes(result.document, result.reconciliation))
    )["Exceptions"]
    headers = [cell.value for cell in sheet[1]]
    assert "Review" in headers and "Reviewer note" in headers
    rules = [sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)]
    assert "R-01" in rules


def test_the_workbook_names_which_rows_were_corrected(broken):
    finding = _finding(broken, "R-01")
    result = resolve_finding(
        broken, finding, ReviewAction.CORRECTED, corrected_value=str(finding.expected)
    )
    sheet = load_workbook(
        io.BytesIO(to_bytes(result.document, result.reconciliation))
    )["Claim Detail"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("Review") + 1
    states = {sheet.cell(row=r, column=column).value for r in range(2, sheet.max_row + 1)}
    assert "as extracted" in states
    assert any(state and state.startswith("corrected:") for state in states)


# --------------------------------------------------------------------------
# The corpus is untouched unless a test corrects something
# --------------------------------------------------------------------------


def test_a_freshly_read_document_has_no_review_history(broken):
    assert broken.document.review_log.entries == []
    summary = summarise_review(broken.reconciliation.findings, broken.document.review_log)
    assert summary.reviewed == 0
    assert summary.outstanding == summary.total


def test_resolving_needs_no_reviewer_identity(broken):
    """There are no accounts yet, and none is invented."""
    result = resolve_finding(broken, _finding(broken, "R-01"), ReviewAction.CONFIRMED)
    assert result.document.review_log.entries[0].reviewer == "local reviewer"


def test_correcting_a_finding_without_a_field_is_refused(broken):
    """A document-wide finding has no cell to correct, and none is guessed.

    Refused out loud, not absorbed. Recording "corrected" against a finding
    that changed nothing puts an act in the audit trail that never happened,
    which is worse than telling the reviewer this is not correctable.
    """
    document_wide = next(
        f for f in broken.reconciliation.findings if not f.claim_number
    )
    before = broken.document.model_dump_json(exclude={"review_log"})
    with pytest.raises(ValueError, match="whole document"):
        resolve_finding(broken, document_wide, ReviewAction.CORRECTED)
    assert broken.document.model_dump_json(exclude={"review_log"}) == before
    assert not broken.document.review_log.entries

    # It is still reviewable, just not correctable.
    result = resolve_finding(broken, document_wide, ReviewAction.CONFIRMED)
    assert result.document.review_log.entries[-1].action is ReviewAction.CONFIRMED
