"""Stage 7 — the exported workbook."""

from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from core.export import (
    DEFAULT_TEMPLATE,
    EXPORT_TEMPLATES,
    build_workbook,
    resolve_columns,
    _DOCUMENT_COLUMNS,
    suggested_filename,
    to_bytes,
    write_xlsx,
)
from core.pipeline import run_pipeline
from core.reconcile import reconcile
from core.schema import REDACTED_FIELDS, Claim, DocumentStatus, LossRunDocument


@pytest.fixture()
def extracted(golden_dir):
    return run_pipeline(golden_dir / "arithmetic_error.pdf", use_vision=False)


def test_sheets_in_order(extracted):
    workbook = build_workbook(extracted.document, extracted.reconciliation)
    assert workbook.sheetnames == ["Claim Detail", "Loss Summary", "Large Loss", "Exceptions", "Source Info"]


def test_claims_sheet_has_a_row_per_claim(extracted):
    sheet = build_workbook(extracted.document, extracted.reconciliation)["Claim Detail"]
    assert sheet.max_row == len(extracted.document.claims) + 1
    assert sheet["A1"].value == "Claim number"
    assert sheet["A2"].value == "FM-0001"


def test_money_cells_carry_a_money_format(extracted):
    sheet = build_workbook(extracted.document, extracted.reconciliation)["Claim Detail"]
    columns = resolve_columns(DEFAULT_TEMPLATE)
    index = columns.index("incurred_total") + 1
    assert "#,##0.00" in sheet.cell(row=2, column=index).number_format


def test_exceptions_sheet_lists_every_finding(extracted):
    sheet = build_workbook(extracted.document, extracted.reconciliation)["Exceptions"]
    assert sheet.max_row == len(extracted.reconciliation.findings) + 1
    assert sheet["A2"].value == "R-01"
    assert sheet["C2"].value == "FM-0003"


def test_exceptions_sheet_says_so_when_clean(golden_dir):
    result = run_pipeline(golden_dir / "us_basic.pdf", use_vision=False)
    sheet = build_workbook(result.document, result.reconciliation)["Exceptions"]
    assert result.reconciliation.status is DocumentStatus.CLEAN
    # The sheet lists the unstated-basis flags rather than claiming silence.
    assert {sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)} == {"R-18"}


def test_source_info_carries_the_audit_trail(extracted):
    sheet = build_workbook(extracted.document, extracted.reconciliation)["Source Info"]
    facts = {row[0].value: row[1].value for row in sheet.iter_rows(min_row=1, max_col=2)}
    assert facts["Source file"] == "arithmetic_error.pdf"
    assert facts["SHA-256"] == extracted.document.file_sha256
    assert facts["Valuation date"] == date(2024, 12, 31)
    assert facts["Extraction method"] == "digital"
    assert facts["Reconciliation status"] == "Needs review"
    assert facts["Claims extracted"] == 4
    assert facts["Exported at"]


def test_source_info_compares_printed_and_extracted_totals(golden_dir):
    result = run_pipeline(golden_dir / "us_basic.pdf", use_vision=False)
    sheet = build_workbook(result.document, result.reconciliation)["Source Info"]
    text = [
        [cell.value for cell in row]
        for row in sheet.iter_rows()
    ]
    flat = [value for row in text for value in row if value is not None]
    assert "Printed totals vs extracted" in flat
    assert 156341.90 in [v for v in flat if isinstance(v, float)]


def test_redaction_drops_personal_data(golden_dir):
    result = run_pipeline(golden_dir / "us_basic.pdf", use_vision=False)
    assert any(claim.claimant_name for claim in result.document.claims)

    workbook = build_workbook(result.document, result.reconciliation,
                              template="Full detail", redact=True)
    headers = [cell.value for cell in workbook["Claim Detail"][1]]
    for field_name in REDACTED_FIELDS:
        assert all(field_name not in (header or "").lower().replace(" ", "_")
                   for header in headers)
    body = "\n".join(
        str(cell.value)
        for row in workbook["Claim Detail"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Alvarez" not in body
    assert "Slip and fall" not in body


def test_redaction_is_off_by_default(golden_dir):
    result = run_pipeline(golden_dir / "us_basic.pdf", use_vision=False)
    workbook = build_workbook(result.document, result.reconciliation, template="Full detail")
    body = "\n".join(
        str(cell.value)
        for row in workbook["Claim Detail"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Alvarez, Marisol" in body


@pytest.mark.parametrize("template", list(EXPORT_TEMPLATES))
def test_every_template_builds(extracted, template):
    workbook = build_workbook(extracted.document, extracted.reconciliation, template=template)
    headers = [cell.value for cell in workbook["Claim Detail"][1]]
    assert headers[0] == "Claim number"
    # Every row also carries the document-level facts, so a row read on its own
    # still says which carrier, policy and valuation date it belongs to.
    assert len(headers) == len(resolve_columns(template)) + len(_DOCUMENT_COLUMNS)
    assert headers[-len(_DOCUMENT_COLUMNS)] == "Carrier"


def test_custom_column_order():
    document = LossRunDocument(
        source_filename="a.pdf", file_sha256="x",
        claims=[Claim(claim_number="C1", incurred_total=Decimal("100"))],
    )
    workbook = build_workbook(document, None, template=["incurred_total", "claim_number"],
                              include_provenance=False)
    headers = [cell.value for cell in workbook["Claim Detail"][1]]
    assert headers[:2] == ["Total incurred", "Claim number"]
    assert headers[2:] == [title for title, _ in _DOCUMENT_COLUMNS]


def test_provenance_columns_can_be_dropped(extracted):
    with_provenance = resolve_columns(DEFAULT_TEMPLATE, include_provenance=True)
    without = resolve_columns(DEFAULT_TEMPLATE, include_provenance=False)
    assert "source_page" in with_provenance and "source_page" not in without


def test_findings_shade_the_cell_they_belong_to(extracted):
    sheet = build_workbook(extracted.document, extracted.reconciliation)["Claim Detail"]
    columns = resolve_columns(DEFAULT_TEMPLATE)
    incurred = columns.index("incurred_total") + 1
    broken_row = next(
        index
        for index, claim in enumerate(extracted.document.claims, start=2)
        if claim.claim_number == "FM-0003"
    )
    assert sheet.cell(row=broken_row, column=incurred).fill.fgColor.rgb.endswith("F8D7DA")
    clean_row = 2 if broken_row != 2 else 3
    assert not sheet.cell(row=clean_row, column=incurred).fill.fgColor.rgb.endswith("F8D7DA")


def test_bytes_round_trip_through_openpyxl(extracted, tmp_path):
    payload = to_bytes(extracted.document, extracted.reconciliation)
    path = tmp_path / "out.xlsx"
    path.write_bytes(payload)
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Claim Detail", "Loss Summary", "Large Loss", "Exceptions", "Source Info"]


def test_write_xlsx_creates_the_file(extracted, tmp_path):
    path = write_xlsx(extracted.document, tmp_path / "nested" / "out.xlsx",
                      extracted.reconciliation)
    assert path.exists() and path.stat().st_size > 0


def test_suggested_filename_is_findable(extracted):
    name = suggested_filename(extracted.document)
    assert name.endswith(".xlsx")
    assert "Copper Creek Brewing Co" in name
    assert "2024-12-31" in name


def test_suggested_filename_flags_a_missing_valuation_date():
    document = LossRunDocument(source_filename="x.pdf", file_sha256="h")
    assert "no-valuation-date" in suggested_filename(document)


def test_export_survives_an_empty_document():
    document = LossRunDocument(source_filename="empty.pdf", file_sha256="h")
    workbook = build_workbook(document, reconcile(document))
    assert workbook["Claim Detail"].max_row == 1
    assert workbook["Exceptions"].max_row >= 2   # R-06 fires: no valuation date
