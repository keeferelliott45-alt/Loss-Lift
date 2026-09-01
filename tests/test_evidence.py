"""Where did this number come from?

An underwriter who cannot answer that about a figure cannot use it, which is
why provenance is principle 2 of the spec rather than a nicety. The chain has
four links and a value has to keep its evidence across all of them: the words
are read off a page, parsed into a value, checked by the rule engine, and
written into a workbook that will outlive the upload.

The harder half of this file is about refusal. A rectangle is recorded in the
space the word extractor reports, which on a rotated page is not the space the
text lives in — so a rectangle can be arithmetically perfect and point at the
wrong part of the page. Marking the wrong row is worse than marking none, so
every region is read back before it is shown, and withdrawn if the claim is not
inside it.
"""

from __future__ import annotations

import dataclasses
from datetime import date
from decimal import Decimal

import pymupdf
import pytest

from core.evidence import (
    EvidenceKind,
    claim_evidence,
    confirm_region,
    finding_evidence,
    render_evidence,
)
from core.export import to_bytes
from core.pipeline import apply_edits, run_pipeline, to_records, review_columns
from core.schema import Claim, Finding, Severity, SourceMethod

FONT = "helv"

COLUMNS = ((30, "Claim Number"), (150, "Date of Loss"), (250, "Status"),
           (330, "Total Paid"), (430, "Total Reserves"), (540, "Total Incurred"))

CLAIMS = (
    ("EV-1001", "3/14/2024", "Closed", "1,200.00", "0.00", "1,200.00"),
    ("EV-1002", "5/02/2024", "Open", "400.00", "900.00", "1,300.00"),
)


#: The page a landscape sheet is really made of: portrait, with the text drawn
#: sideways and ``/Rotate`` turning the whole thing to read across. AIG's pages
#: are built exactly this way, and it is the case where a rectangle recorded in
#: one coordinate space and read in another lands somewhere else entirely.
PORTRAIT = (612, 792)


def _build(path, *, rotated: bool = False) -> None:
    document = pymupdf.open()
    if not rotated:
        page = document.new_page(width=700, height=300)
        place = lambda x, y: (x, y)  # noqa: E731
        turn = 0
    else:
        page = document.new_page(width=PORTRAIT[0], height=PORTRAIT[1])
        # Drawn sideways, so that /Rotate 270 brings it upright: a value meant
        # for (x, top) of the turned page is written at (height - top, x).
        place = lambda x, y: (PORTRAIT[0] - y, x)  # noqa: E731
        turn = 270

    def write(text: str, x: float, y: float, size: float = 8.0) -> None:
        page.insert_text(place(x, y), text, fontname=FONT, fontsize=size, rotate=turn)

    write("Fairmount Underwriters", 30, 30, 9)
    write("Valuation Date: 12/31/2024", 400, 30)
    for x, label in COLUMNS:
        write(label, x, 62)
    y = 90
    for row in CLAIMS:
        for (x, _), value in zip(COLUMNS, row):
            write(value, x, y)
        y += 18
    if rotated:
        page.set_rotation(turn)
    document.save(str(path))
    document.close()


@pytest.fixture(scope="module")
def upright(tmp_path_factory):
    path = tmp_path_factory.mktemp("evidence") / "upright.pdf"
    _build(path)
    return run_pipeline(path, use_vision=False)


@pytest.fixture(scope="module")
def rotated(tmp_path_factory):
    """The same page turned ninety degrees, as landscape sheets arrive."""
    path = tmp_path_factory.mktemp("evidence") / "rotated.pdf"
    _build(path, rotated=True)
    return run_pipeline(path, use_vision=False)


# --------------------------------------------------------------------------
# The chain: extraction -> normalisation -> reconciliation -> export
# --------------------------------------------------------------------------


def test_a_digital_claim_keeps_its_page_line_and_region(upright):
    claim = upright.document.claims[0]
    assert claim.source_page == 1
    assert claim.source_row is not None
    assert claim.source_bbox is not None
    assert claim.source_method is SourceMethod.DIGITAL


def test_the_region_survives_being_read_back(upright):
    """The whole point: the marked area really does contain this claim."""
    claim = upright.document.claims[0]
    evidence = claim_evidence(claim)
    assert evidence.kind is EvidenceKind.REGION
    confirmed = confirm_region(upright.source_path, evidence, claim.claim_number)
    assert confirmed.kind is EvidenceKind.REGION
    assert confirmed.bbox == claim.source_bbox


def test_a_field_carries_the_text_it_was_read_from(upright):
    """Not just where on the page — what the page actually said there."""
    claim = next(c for c in upright.document.claims if c.claim_number == "EV-1002")
    evidence = claim_evidence(claim, "reserve_total")
    assert evidence.text == "900.00"
    assert claim.reserve_total == Decimal("900.00")


def test_a_finding_points_at_the_claim_it_is_about(upright):
    """A reviewer clicking an exception should land on the row that caused it."""
    document = upright.document
    claim = document.claims[0]
    finding = Finding(
        rule_id="R-01", severity=Severity.ERROR, message="…",
        claim_number=claim.claim_number, field="incurred_total",
    )
    evidence = finding_evidence(document, finding)
    assert evidence.page == claim.source_page
    assert evidence.bbox == claim.source_bbox
    assert evidence.kind is EvidenceKind.REGION


def test_a_document_wide_finding_does_not_pretend_to_a_row(upright):
    """R-06 is about the document. There is no row to mark, so none is."""
    finding = Finding(rule_id="R-06", severity=Severity.ERROR, message="…")
    evidence = finding_evidence(upright.document, finding)
    assert evidence.kind is EvidenceKind.NONE
    assert evidence.bbox is None


def _source_location(document, **kwargs) -> str:
    import io

    from openpyxl import load_workbook

    data = to_bytes(
        document,
        template=["claim_number", "incurred_total"],
        include_provenance=True,
        **kwargs,
    )
    sheet = load_workbook(io.BytesIO(data))["Claim Detail"]
    headers = [cell.value for cell in sheet[1]]
    assert "Source location" in headers
    assert "Extraction" in headers
    return sheet.cell(row=2, column=headers.index("Source location") + 1).value


def test_provenance_reaches_the_workbook(upright):
    """The export outlives the upload, so it carries the evidence in words.

    A region is written down only when it has been read back, which needs the
    file. Given it, the workbook names the rectangle; without it, the workbook
    names the page and stops — the same degradation the review screen makes,
    for the same reason. Guessing in the direction of precision is how a
    filed workbook ends up pointing at a row nobody checked.
    """
    stated = _source_location(upright.document, source_path=upright.source_path)
    assert stated.startswith("page 1")
    assert "at (" in stated

    unverified = _source_location(upright.document)
    assert unverified.startswith("page 1")
    assert "at (" not in unverified


def test_provenance_reaches_the_review_table(upright):
    records = to_records(upright.document, review_columns(upright.document))
    assert records[0]["_page"] == 1
    assert records[0]["_method"] == "digital"
    assert records[0]["_row"] is not None


# --------------------------------------------------------------------------
# Rotated pages: the same rectangle, a different space
# --------------------------------------------------------------------------


def test_a_rotated_page_is_marked_in_the_right_place(rotated):
    claim = rotated.document.claims[0]
    evidence = claim_evidence(claim)
    confirmed = confirm_region(rotated.source_path, evidence, claim.claim_number)
    assert confirmed.kind is EvidenceKind.REGION, confirmed.note


def test_the_highlight_is_drawn_where_the_claim_is(rotated):
    """Rendered with and without the mark, only the claim's row differs."""
    claim = rotated.document.claims[0]
    evidence = confirm_region(
        rotated.source_path, claim_evidence(claim), claim.claim_number
    )
    plain = dataclasses.replace(evidence, kind=EvidenceKind.PAGE, bbox=None)
    marked = render_evidence(rotated.source_path, evidence, dpi=72)
    bare = render_evidence(rotated.source_path, plain, dpi=72)
    assert marked is not None and bare is not None and marked != bare

    changed = _changed_box(marked, bare)
    x0, top, x1, bottom = evidence.bbox
    assert changed is not None
    assert abs(changed[0] - x0) < 8 and abs(changed[1] - top) < 8
    assert abs(changed[2] - x1) < 8 and abs(changed[3] - bottom) < 8


def _changed_box(png_a: bytes, png_b: bytes):
    """The bounding box of every pixel the highlight changed.

    Compared a row of bytes at a time: the two renders are identical except
    where the mark was drawn, so almost every row is skipped whole.
    """
    import io

    left = pymupdf.Pixmap(io.BytesIO(png_a))
    right = pymupdf.Pixmap(io.BytesIO(png_b))
    a, b, stride, n = left.samples, right.samples, left.stride, left.n
    x0 = y0 = None
    x1 = y1 = -1
    for y in range(left.height):
        start = y * stride
        row_a = a[start : start + stride]
        row_b = b[start : start + stride]
        if row_a == row_b:
            continue
        differing = [x for x in range(left.width)
                     if row_a[x * n : x * n + n] != row_b[x * n : x * n + n]]
        if not differing:
            continue
        y0 = y if y0 is None else y0
        y1 = y
        x0 = min(differing) if x0 is None else min(x0, min(differing))
        x1 = max(x1, max(differing))
    return (x0, y0, x1, y1) if x0 is not None else None


# --------------------------------------------------------------------------
# Refusal
# --------------------------------------------------------------------------


def test_a_region_that_does_not_hold_the_claim_is_withdrawn(upright):
    """The protection against marking a plausible but wrong row."""
    claim = upright.document.claims[0]
    wrong = dataclasses.replace(claim_evidence(claim), bbox=(500.0, 250.0, 600.0, 270.0))
    confirmed = confirm_region(upright.source_path, wrong, claim.claim_number)
    assert confirmed.kind is EvidenceKind.PAGE
    assert confirmed.bbox is None
    assert "worse than marking none" in confirmed.note


def test_a_deleted_upload_costs_the_region_but_not_the_record(upright):
    """After export the file is gone; the page, line and text remain."""
    claim = upright.document.claims[0]
    evidence = claim_evidence(claim, "incurred_total")
    confirmed = confirm_region("/no/such/file.pdf", evidence, claim.claim_number)
    assert confirmed.kind is EvidenceKind.PAGE
    assert confirmed.page == claim.source_page
    assert confirmed.text == evidence.text
    assert render_evidence("/no/such/file.pdf", confirmed) is None


def test_a_vision_claim_offers_the_page_and_no_rectangle():
    """The model is asked what the page says, not where the words sit."""
    claim = Claim(
        claim_number="V-1", date_of_loss=date(2024, 5, 1),
        source_page=9, source_method=SourceMethod.VISION,
        raw_cells={"incurred_total": "1,000.00"},
    )
    evidence = claim_evidence(claim, "incurred_total")
    assert evidence.kind is EvidenceKind.PAGE
    assert evidence.method is SourceMethod.VISION
    assert evidence.bbox is None
    assert evidence.page == 9
    assert evidence.text == "1,000.00"
    assert "vision model" in evidence.note
    assert evidence.describe() == "page 9"


# --------------------------------------------------------------------------
# Manual edits
# --------------------------------------------------------------------------


def test_an_edited_field_says_a_person_typed_it(upright):
    records = to_records(upright.document, review_columns(upright.document))
    records[0]["incurred_total"] = 9999.0
    updated = apply_edits(upright.document, records)
    claim = updated.claims[0]
    assert "incurred_total" in claim.edited_fields
    evidence = claim_evidence(claim, "incurred_total")
    assert evidence.kind is EvidenceKind.TYPED
    assert evidence.method is SourceMethod.MANUAL
    assert evidence.describe() == "entered by hand"


def test_editing_one_cell_does_not_cost_the_others_their_provenance(upright):
    """Nine untouched figures still came off the page, and still say so."""
    records = to_records(upright.document, review_columns(upright.document))
    records[0]["incurred_total"] = 9999.0
    claim = apply_edits(upright.document, records).claims[0]
    untouched = claim_evidence(claim, "paid_total")
    assert untouched.kind is EvidenceKind.REGION
    assert untouched.bbox == upright.document.claims[0].source_bbox
    assert untouched.text == "1,200.00"
    # The claim as a whole is marked manual once any cell is corrected, so the
    # note has to say which cell that was — otherwise a reviewer looking at
    # this figure cannot tell whether it is one of them.
    assert "Incurred total" in untouched.note


def test_a_claim_added_by_hand_claims_no_source():
    claim = Claim(claim_number="NEW-1", source_method=SourceMethod.MANUAL)
    evidence = claim_evidence(claim)
    assert evidence.kind is EvidenceKind.TYPED
    assert "not read from the document" in evidence.note


# --------------------------------------------------------------------------
# Multi-line records
# --------------------------------------------------------------------------


def test_a_record_spread_over_several_lines_names_all_of_them():
    claim = Claim(
        claim_number="ML-1", source_page=2, source_row=9,
        source_lines=[8, 9, 10], source_bbox=(7.0, 179.0, 769.0, 202.0),
    )
    evidence = claim_evidence(claim)
    assert evidence.lines == [8, 9, 10]
    assert "lines 9, 10, 11" in evidence.describe()


def test_a_single_line_record_says_line_not_lines():
    claim = Claim(claim_number="SL-1", source_page=1, source_row=4)
    assert "line 5" in claim_evidence(claim).describe()

# --------------------------------------------------------------------------
# The two libraries do not describe a page the same way
# --------------------------------------------------------------------------


@pytest.fixture(scope="module")
def cropped(tmp_path_factory):
    """A page whose crop box starts a little way down and across.

    The word extractor measures from the media box and PyMuPDF from the crop
    box, so on a document like this the same rectangle names two different
    parts of the page — about a line apart, which is enough to mark the row
    below the right one. Seventeen of Liberty's twenty-eight claims were
    marked that way before the offset was accounted for.
    """
    plain = tmp_path_factory.mktemp("evidence") / "plain.pdf"
    _build(plain)
    path = tmp_path_factory.mktemp("evidence") / "cropped.pdf"
    document = pymupdf.open(plain)
    document[0].set_cropbox(pymupdf.Rect(7, 9, 700, 300))
    document.save(str(path))
    document.close()
    return run_pipeline(path, use_vision=False)


def test_a_shifted_crop_box_still_marks_the_right_row(cropped):
    claim = cropped.document.claims[0]
    evidence = claim_evidence(claim)
    assert evidence.kind is EvidenceKind.REGION
    confirmed = confirm_region(cropped.source_path, evidence, claim.claim_number)
    assert confirmed.kind is EvidenceKind.REGION, confirmed.note


def test_every_claim_on_a_cropped_page_is_locatable(cropped):
    for claim in cropped.document.claims:
        confirmed = confirm_region(
            cropped.source_path, claim_evidence(claim), claim.claim_number
        )
        assert confirmed.kind is EvidenceKind.REGION, (
            f"{claim.claim_number}: {confirmed.note}"
        )

def test_a_vision_claim_edited_in_one_cell_keeps_vision_for_the_rest():
    """The contract, on the path most likely to get it wrong.

    A scanned claim corrected in one cell is manual there and still read by
    the vision model everywhere else. Reporting the whole row as manual would
    hide that the other nine figures came out of an image.
    """
    claim = Claim(
        claim_number="V-2", source_page=9,
        source_method=SourceMethod.MANUAL, read_method=SourceMethod.VISION,
        edited_fields=["paid_total"],
        raw_cells={"paid_total": "100.00", "reserve_total": "50.00"},
        original_values={"paid_total": "10.00"},
    )
    assert claim.provenance_of("paid_total") is SourceMethod.MANUAL
    assert claim.provenance_of("reserve_total") is SourceMethod.VISION

    corrected = claim_evidence(claim, "paid_total")
    assert corrected.kind is EvidenceKind.TYPED

    untouched = claim_evidence(claim, "reserve_total")
    assert untouched.kind is EvidenceKind.PAGE
    assert untouched.method is SourceMethod.VISION
    assert untouched.bbox is None
    assert "vision model" in untouched.note
