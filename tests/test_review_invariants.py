"""Eight things that must hold however a document is worked on.

The tests next door each reproduce one way the review workflow went wrong. These
are the other direction: eight properties stated once, then checked after every
kind of operation a reviewer can perform — confirming, dismissing, correcting,
editing a cell in the table, renaming a claim, adding a row, deleting a row, and
undoing an edit.

Written this way on purpose. Every defect this file guards against was found by
combining two operations that were each fine alone: a correction after a rename,
a dismissal after an edit. A property that only holds for the operation someone
thought to test is not a property.

    1. Carrier evidence is never lost.
    2. A decision attaches to one finding, and to the right one.
    3. A document never reads healthier than the checks run against it.
    4. Audit history is appended, never rewritten or dropped.
    5. Changing a value re-runs every rule, or changes nothing.
    6. The workbook says what the screen says.
    7. A typed value never reads as the carrier's.
    8. A decision stops applying when its finding materially changes.
"""

from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from core.evidence import EvidenceKind, claim_evidence, confirm_region
from core.export import _cell_value, to_bytes
from core.pipeline import (
    apply_edits,
    rerun_reconciliation,
    resolve_finding,
    review_columns,
    run_pipeline,
    to_records,
)
from core.reconcile import reconcile
from core.review import ReviewAction, finding_key, summarise_review
from core.schema import SourceMethod


@pytest.fixture()
def broken(golden_dir):
    return run_pipeline(golden_dir / "arithmetic_error.pdf", use_vision=False)


# --------------------------------------------------------------------------
# Every way a reviewer can change a document
# --------------------------------------------------------------------------


def _grid(result, mutate):
    records = to_records(result.document, review_columns(result.document))
    mutate(records)
    result.document = apply_edits(result.document, records)
    result.reconciliation = rerun_reconciliation(result.document)
    return result


def _confirm_first(result):
    return resolve_finding(result, result.reconciliation.findings[0], ReviewAction.CONFIRMED)


def _dismiss_everything(result):
    for finding in list(result.reconciliation.findings):
        result = resolve_finding(result, finding, ReviewAction.DISMISSED, note="looked")
    return result


def _correct_the_error(result):
    failing = next(f for f in result.reconciliation.findings if f.rule_id == "R-01")
    return resolve_finding(result, failing, ReviewAction.CORRECTED, corrected_value="31400.00")


def _edit_a_cell(result):
    return _grid(result, lambda rows: rows[0].__setitem__("incurred_total", 12345.0))


def _rename_a_claim(result):
    return _grid(result, lambda rows: rows[0].__setitem__("claim_number", "RENUMBERED-1"))


def _blank_a_cell(result):
    return _grid(result, lambda rows: rows[0].__setitem__("paid_total", ""))


def _delete_a_row(result):
    def drop(rows):
        rows.pop(1)
    return _grid(result, drop)


def _add_a_row(result):
    def append(rows):
        blank = {name: None for name in rows[0]}
        blank["claim_number"] = "TYPED-1"
        blank["incurred_total"] = 500.0
        rows.append(blank)
    return _grid(result, append)


def _undo_an_edit(result):
    was = result.document.claims[0].incurred_total
    result = _grid(result, lambda rows: rows[0].__setitem__("incurred_total", 7.0))
    return _grid(result, lambda rows: rows[0].__setitem__("incurred_total", float(was)))


def _correct_then_dismiss(result):
    return _dismiss_everything(_correct_the_error(result))


def _rename_then_correct(result):
    result = _rename_a_claim(result)
    failing = next(f for f in result.reconciliation.findings if f.rule_id == "R-01")
    return resolve_finding(result, failing, ReviewAction.CORRECTED, corrected_value="31400.00")


OPERATIONS = {
    "confirm": _confirm_first,
    "dismiss everything": _dismiss_everything,
    "correct against a finding": _correct_the_error,
    "edit a cell": _edit_a_cell,
    "rename a claim": _rename_a_claim,
    "blank a cell": _blank_a_cell,
    "delete a row": _delete_a_row,
    "add a row": _add_a_row,
    "undo an edit": _undo_an_edit,
    "correct then dismiss": _correct_then_dismiss,
    "rename then correct": _rename_then_correct,
}

every_operation = pytest.mark.parametrize("operation", sorted(OPERATIONS), ids=str)


# --------------------------------------------------------------------------
# 1. Carrier evidence is never lost
# --------------------------------------------------------------------------


@every_operation
def test_carrier_evidence_survives(broken, operation):
    """What the document printed stays readable, whatever happens on top of it.

    The printed text of every cell, the region it was read from, the lines it
    was assembled out of, and the reason any null was null. A claim that leaves
    entirely leaves a record of what it held.
    """
    before = {c.row_id: c for c in broken.document.claims}
    result = OPERATIONS[operation](broken)
    after = {c.row_id: c for c in result.document.claims}

    for row_id, was in before.items():
        now = after.get(row_id)
        if now is None:
            deleted = [
                e for e in result.document.review_log.entries
                if e.row_id == row_id and e.action is ReviewAction.DELETED
            ]
            assert deleted, f"{row_id} vanished with no record"
            assert str(was.incurred_total) in (deleted[-1].before or "")
            continue

        assert now.raw_cells == was.raw_cells, f"{row_id} lost the printed text"
        assert now.source_bbox == was.source_bbox, f"{row_id} lost its region"
        assert now.source_lines == was.source_lines
        assert now.source_page == was.source_page
        assert (now.read_method or now.source_method) == (
            was.read_method or was.source_method
        ), f"{row_id} forgot how it was read"
        for field, reason in was.field_issues.items():
            assert (
                now.field_issues.get(field) is reason
                or now.original_issue_of(field) is reason
            ), f"{row_id} lost why {field} was null"


@every_operation
def test_a_changed_value_keeps_the_one_it_replaced(broken, operation):
    before = {c.row_id: c for c in broken.document.claims}
    result = OPERATIONS[operation](broken)
    for claim in result.document.claims:
        was = before.get(claim.row_id)
        if was is None:
            continue
        for field in claim.edited_fields:
            if field in was.edited_fields:
                continue  # already carried an original before this operation
            assert claim.original_of(field) is not None, (
                f"{claim.row_id}.{field} was overwritten with nothing kept"
            )


# --------------------------------------------------------------------------
# 2. A decision attaches to one finding, and to the right one
# --------------------------------------------------------------------------


@every_operation
def test_no_two_findings_share_an_identity(broken, operation):
    result = OPERATIONS[operation](broken)
    keys = [finding_key(f) for f in result.reconciliation.findings]
    assert len(set(keys)) == len(keys), f"one decision would answer for several: {keys}"


def test_a_decision_resolves_the_finding_it_was_taken_about_and_no_other(broken):
    for finding in list(broken.reconciliation.findings):
        result = resolve_finding(broken, finding, ReviewAction.DISMISSED)
        log = result.document.review_log
        assert log.is_resolved(finding)
        others = [f for f in result.reconciliation.findings if finding_key(f) != finding_key(finding)]
        resolved_others = [f for f in others if log.is_resolved(f)]
        assert not resolved_others, (
            f"dismissing {finding_key(finding)} also answered for "
            f"{[finding_key(f) for f in resolved_others]}"
        )
        break


@every_operation
def test_a_decision_names_the_row_it_was_about(broken, operation):
    result = OPERATIONS[operation](broken)
    rows = {c.row_id for c in result.document.claims}
    for entry in result.document.review_log.entries:
        if not entry.row_id:
            assert not entry.claim_number, "a claim-level decision must name its row"
            continue
        assert entry.where, f"{entry.row_id} was recorded with no place on the page"
        if entry.action is not ReviewAction.DELETED:
            assert entry.row_id in rows


# --------------------------------------------------------------------------
# 3. A document never reads healthier than the checks run against it
# --------------------------------------------------------------------------


@every_operation
def test_review_never_makes_a_bucket_pass(broken, operation):
    result = OPERATIONS[operation](broken)
    summary = summarise_review(result.reconciliation.findings, result.document.review_log)
    for name in ("financial", "extraction", "underwriting"):
        bucket = getattr(summary, name)
        assert bucket.passes == (not bucket.findings), (
            f"{name} passes on {bucket.total} findings"
        )


@every_operation
def test_the_headline_names_the_worst_thing_that_is_true(broken, operation):
    result = OPERATIONS[operation](broken)
    summary = summarise_review(result.reconciliation.findings, result.document.review_log)
    headline = summary.headline()
    if not summary.extraction.passes:
        assert headline == "not read cleanly"
    elif not summary.financial.passes:
        assert headline == "does not reconcile"
    elif summary.underwriting.findings:
        assert headline.startswith("reconciled, flags")
    else:
        assert headline == "reconciled"


def test_dismissing_everything_on_a_failing_document_changes_no_status(broken):
    before = broken.reconciliation.status
    findings = [(f.rule_id, f.severity, finding_key(f)) for f in broken.reconciliation.findings]
    result = _dismiss_everything(broken)
    assert result.reconciliation.status is before
    assert [
        (f.rule_id, f.severity, finding_key(f)) for f in result.reconciliation.findings
    ] == findings
    assert summarise_review(
        result.reconciliation.findings, result.document.review_log
    ).headline() == "does not reconcile"


# --------------------------------------------------------------------------
# 4. Audit history is appended, never rewritten or dropped
# --------------------------------------------------------------------------


@every_operation
def test_history_only_grows(broken, operation):
    seeded = resolve_finding(
        broken, broken.reconciliation.findings[0], ReviewAction.CONFIRMED, note="first"
    )
    before = [e.model_dump(mode="json") for e in seeded.document.review_log.entries]
    result = OPERATIONS[operation](seeded)
    after = [e.model_dump(mode="json") for e in result.document.review_log.entries]
    assert after[: len(before)] == before, "an earlier decision was rewritten"
    assert len(after) >= len(before)


@every_operation
def test_every_value_change_is_recorded(broken, operation):
    before = {c.row_id: c for c in broken.document.claims}
    result = OPERATIONS[operation](broken)
    logged = {
        (e.row_id, e.field) for e in result.document.review_log.entries if e.changed_a_value
    }
    logged_rows = {e.row_id for e in result.document.review_log.entries}

    for claim in result.document.claims:
        was = before.get(claim.row_id)
        if was is None:
            assert claim.row_id in logged_rows, f"{claim.row_id} appeared unrecorded"
            continue
        for field in claim.edited_fields:
            if getattr(claim, field) == getattr(was, field, None):
                continue
            assert (claim.row_id, field) in logged, (
                f"{claim.row_id}.{field} changed with nothing in the log"
            )

    for row_id in before:
        if row_id not in {c.row_id for c in result.document.claims}:
            assert row_id in logged_rows, f"{row_id} was deleted unrecorded"


# --------------------------------------------------------------------------
# 5. Changing a value re-runs every rule, or changes nothing
# --------------------------------------------------------------------------


@every_operation
def test_the_findings_match_the_claims_they_are_about(broken, operation):
    """Whatever happened, the findings on screen are the ones the rules give now."""
    result = OPERATIONS[operation](broken)
    fresh = reconcile(result.document)
    assert [finding_key(f) for f in fresh.findings] == [
        finding_key(f) for f in result.reconciliation.findings
    ]
    assert fresh.status is result.reconciliation.status


def test_a_failed_rerun_leaves_the_document_alone(broken, monkeypatch):
    import core.pipeline as pipeline

    failing = next(f for f in broken.reconciliation.findings if f.rule_id == "R-01")
    values = [c.incurred_total for c in broken.document.claims]
    reconciliation = broken.reconciliation

    monkeypatch.setattr(
        pipeline, "rerun_reconciliation",
        lambda *a, **k: (_ for _ in ()).throw(RuntimeError("engine down")),
    )
    with pytest.raises(RuntimeError):
        resolve_finding(broken, failing, ReviewAction.CORRECTED, corrected_value="1.00")

    assert [c.incurred_total for c in broken.document.claims] == values
    assert broken.reconciliation is reconciliation
    assert not broken.document.review_log.entries


# --------------------------------------------------------------------------
# 6. The workbook says what the screen says
# --------------------------------------------------------------------------


@every_operation
def test_the_workbook_matches_the_screen(broken, operation):
    result = OPERATIONS[operation](broken)
    book = load_workbook(
        io.BytesIO(
            to_bytes(
                result.document,
                result.reconciliation,
                template=["claim_number", "incurred_total"],
                source_path=result.source_path,
            )
        )
    )

    sheet = book["Claim Detail"]
    headers = [c.value for c in sheet[1]]
    written = {
        sheet.cell(row=r, column=headers.index("Claim number") + 1).value: sheet.cell(
            row=r, column=headers.index("Total incurred") + 1
        ).value
        for r in range(2, sheet.max_row + 1)
    }
    onscreen = {
        c.claim_number: (None if c.incurred_total is None else float(c.incurred_total))
        for c in result.document.claims
    }
    assert written == onscreen

    exceptions = book["Exceptions"]
    assert exceptions.max_row - 1 == len(result.reconciliation.findings)
    history = book["Review History"]
    entries = result.document.review_log.entries
    # An empty log writes one line saying so, rather than an empty sheet.
    assert history.max_row - 1 == (len(entries) or 1)
    for index, entry in enumerate(entries, start=2):
        assert history.cell(row=index, column=9).value == entry.action.value


@every_operation
def test_the_workbook_never_points_more_precisely_than_the_screen(broken, operation):
    result = OPERATIONS[operation](broken)
    for claim in result.document.claims:
        onscreen = claim_evidence(claim)
        if onscreen.kind is EvidenceKind.REGION:
            onscreen = confirm_region(result.source_path, onscreen, claim.claim_number)
        written = _cell_value(claim, "source_evidence", result.source_path)
        assert written == onscreen.describe()
        if onscreen.kind is not EvidenceKind.REGION:
            assert "at (" not in written


# --------------------------------------------------------------------------
# 7. A typed value never reads as the carrier's
# --------------------------------------------------------------------------


@every_operation
def test_typed_values_say_so(broken, operation):
    result = OPERATIONS[operation](broken)
    for claim in result.document.claims:
        for field in claim.edited_fields:
            assert claim.provenance_of(field) is SourceMethod.MANUAL
            assert claim_evidence(claim, field).kind is EvidenceKind.TYPED
        untouched = [
            f for f in ("paid_total", "reserve_total", "incurred_total")
            if f not in claim.edited_fields
        ]
        if claim.read_method is not None:
            for field in untouched:
                assert claim.provenance_of(field) is claim.read_method


def test_a_row_added_by_hand_does_not_claim_the_document_as_its_source(broken):
    result = _add_a_row(broken)
    added = next(c for c in result.document.claims if c.claim_number == "TYPED-1")
    evidence = claim_evidence(added)
    assert evidence.kind is EvidenceKind.TYPED
    assert "not read from the document" in evidence.note
    assert added.where() == "added on the review screen"
    assert _cell_value(added, "source_evidence") == "entered by hand"


# --------------------------------------------------------------------------
# 8. A decision stops applying when its finding materially changes
# --------------------------------------------------------------------------


@every_operation
def test_a_resolution_only_covers_what_it_was_taken_about(broken, operation):
    seeded = broken
    for finding in list(broken.reconciliation.findings):
        seeded = resolve_finding(seeded, finding, ReviewAction.CONFIRMED, note="seen")
    recorded = {e.key: e for e in seeded.document.review_log.entries}

    result = OPERATIONS[operation](seeded)
    log = result.document.review_log
    for finding in result.reconciliation.findings:
        entry = recorded.get(finding_key(finding))
        if entry is None:
            assert not log.is_resolved(finding), "a finding nobody saw reads as reviewed"
            continue
        moved = (entry.expected, entry.actual, entry.delta) != (
            None if finding.expected is None else str(finding.expected),
            None if finding.actual is None else str(finding.actual),
            None if finding.delta is None else str(finding.delta),
        )
        if moved and not log.is_resolved(finding):
            continue  # retired, which is the point
        assert log.is_resolved(finding) == (not moved) or not moved


def test_correcting_a_figure_retires_the_decision_taken_about_the_old_one(broken):
    failing = next(f for f in broken.reconciliation.findings if f.rule_id == "R-01")
    result = resolve_finding(broken, failing, ReviewAction.CONFIRMED, note="carrier is right")
    result = _grid(
        result,
        lambda rows: rows[2].__setitem__(
            "paid_total", float(Decimal(str(rows[2]["paid_total"] or 0)) + Decimal("500"))
        ),
    )
    now = next(
        (f for f in result.reconciliation.findings
         if f.rule_id == "R-01" and f.subject == failing.subject),
        None,
    )
    if now is None:
        return  # the edit cleared it; nothing left to carry a stale decision
    if now.delta != failing.delta:
        assert not result.document.review_log.is_resolved(now)
    assert any(e.action is ReviewAction.CONFIRMED for e in result.document.review_log.entries)
